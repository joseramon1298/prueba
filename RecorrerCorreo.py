import pymssql
from datetime import datetime, timedelta
import openpyxl
import schedule
import time
import configparser

# Leer los datos de configuración del archivo config.ini
config = configparser.ConfigParser()
config.read('config.yml')

# Parámetros de conexión a la base de datos
server = config.get('Database', 'server')
database = config.get('Database', 'database')
username = config.get('Database', 'username')
password = config.get('Database', 'password')

# Ruta del archivo Excel
excel_file = 'correos.xlsx'

# Obtener el número de días del archivo de configuración
days = int(config.get('Settings', 'days'))

def obtener_correos():
    # Conexión a la base de datos
    conn = pymssql.connect(server=server, database=database, user=username, password=password)

    # Cálculo de la fecha límite
    limite_fecha = datetime.now().date() + timedelta(days=days)

    # Consulta SQL para obtener los registros que cumplen la condición
    sql_query = f"SELECT Email FROM dbo.Correos WHERE Fecha = '{limite_fecha}'"

    # Ejecución de la consulta
    cursor = conn.cursor()
    cursor.execute(sql_query)

    # Obtener los resultados de la consulta
    resultados = cursor.fetchall()

    conn.close()

    return resultados

def escribir_en_excel(correos):
    # Cargar el archivo Excel existente o crear uno nuevo
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Email"])

    # Borrar los datos existentes en el archivo Excel
    sheet.delete_rows(2, sheet.max_row)

    # Escribir los correos en el archivo Excel
    for correo in correos:
        sheet.append([correo[0]])

    # Guardar los cambios en el archivo Excel
    workbook.save(excel_file)

def ejecutar_tarea():
    correos = obtener_correos()
    escribir_en_excel(correos)

# Ejecutar la tarea por primera vez
ejecutar_tarea()

# Programar la ejecución de la tarea cada 10 minutos
schedule.every(10).minutes.do(ejecutar_tarea)

while True:
    schedule.run_pending()
    time.sleep(1)
